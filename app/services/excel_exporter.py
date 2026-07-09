import io
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.models.transaction import Transaction

class ExcelExporter:
    """
    ExcelExporter encapsulates the process of generating XLSX reports from transaction lists.
    """
    def __init__(self):
        pass

    def export_transactions(self, username: str, transactions: List[Transaction], budget_progress: dict) -> io.BytesIO:
        """
        Creates an Excel workbook containing:
        - Sheet 1: Log Transaksi Lengkap
        - Sheet 2: Ringkasan Anggaran Kategori
        Returns raw bytes in an in-memory buffer.
        """
        wb = Workbook()
        
        # Setup Sheet 1: Log Transaksi
        ws1 = wb.active
        ws1.title = "Log Transaksi"
        
        # Headers styling
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        headers_1 = ["No", "Tanggal", "Item/Nama Transaksi", "Kategori", "Tipe", "Nominal"]
        ws1.append(headers_1)
        for col_num in range(1, 7):
            cell = ws1.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Add data rows
        for idx, tx in enumerate(transactions, start=1):
            row_data = [
                idx,
                tx.transaction_date.strftime("%Y-%m-%d") if tx.transaction_date else "",
                tx.item_name,
                tx.category,
                "Pemasukan" if tx.type == "income" else "Pengeluaran",
                tx.amount
            ]
            ws1.append(row_data)

        # Auto-adjust column width
        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # Setup Sheet 2: Ringkasan Kategori
        ws2 = wb.create_sheet(title="Ringkasan Kategori")
        headers_2 = ["Kategori", "Total Pengeluaran", "Limit Anggaran", "Persentase (%)"]
        ws2.append(headers_2)
        for col_num in range(1, 5):
            cell = ws2.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for cat, data in budget_progress.items():
            row_data = [
                cat.capitalize(),
                data["spent"],
                data["limit"],
                data["percentage"]
            ]
            ws2.append(row_data)

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Save to memory stream
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
